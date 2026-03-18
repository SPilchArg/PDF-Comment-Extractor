import re
import threading
import zipfile
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import fitz
import openpyxl
from openpyxl.styles import Alignment
from langdetect import detect, LangDetectException
from collections import Counter
import tempfile
import shutil

ARGOS = "argos"

SKIP_DESCRIPTIONS = {
    "sticky note", "underline", "highlight", "strikeout", "squiggly",
    "ink", "line", "square", "circle", "polygon", "polyline", "stamp",
    "caret", "free text", "freetext", "popup", "file attachment", "sound",
    "movie", "widget", "screen", "printer mark", "trap net", "watermark",
    "3d", "redact",
}

def extract_comments_from_pdf(pdf_path: str,
                               progress_callback=None,
                               status_callback=None) -> list[dict]:
    comments   = []
    comment_id = 1
    file_name  = Path(pdf_path).name
    skip_lower = {s.lower() for s in SKIP_DESCRIPTIONS}

    p = Path(pdf_path)
    if not p.exists():
        raise FileNotFoundError(f"PDF not found on disk: {pdf_path}")
    if p.stat().st_size == 0:
        raise ValueError(f"PDF is empty (0 bytes): {pdf_path}")

    doc         = fitz.open(pdf_path)
    total_pages = len(doc)

    if status_callback:
        status_callback(f"  Opened '{file_name}' — {total_pages} page(s)")

    for page_num in range(total_pages):
        page           = doc[page_num]
        full_page_text = page.get_text()

        if status_callback:
            status_callback(f"  Page {page_num + 1}/{total_pages} …")
        if progress_callback:
            progress_callback((page_num + 1) / total_pages * 100)

        for annot in page.annots():
            raw_text = _get_annotation_text(annot)
            if not raw_text:
                continue
            if raw_text.strip().lower() in skip_lower:
                continue

            parsed      = parse_annotation_text(raw_text)
            description = format_description(parsed)

            if _is_junk_annotation(description, skip_lower):
                continue

            keywords       = build_keywords(parsed)
            annotated_text = _get_annotated_text(annot, page, full_page_text)
            language       = _detect_language(annotated_text)

            comments.append({
                "no":            comment_id,
                "file_name":     file_name,
                "description":   description,
                "found_on_page": page_num + 1,
                "language":      language,
                "keywords":      keywords,
            })
            comment_id += 1

    doc.close()
    return comments


def _get_annotation_text(annot) -> str:
    content = (annot.info.get("content") or "").strip()
    if not content:
        try:
            content = (annot.get_text() or "").strip()
        except Exception:
            pass
    return content


def _get_annotated_text(annot, page, full_page_text: str) -> str:
    rect = annot.rect
    for pad in (200, 400, 800):
        try:
            clip = fitz.Rect(
                max(rect.x0 - pad, 0), max(rect.y0 - pad, 0),
                rect.x1 + pad,         rect.y1 + pad,
            )
            text = page.get_text("text", clip=clip).strip()
            if len(text) >= 80:
                return text
        except Exception:
            pass
    return full_page_text


def _is_junk_annotation(description: str, skip_lower: set) -> bool:
    return description.strip().lower() in skip_lower


def parse_annotation_text(text: str) -> dict:
    fields = {
        "author": "", "status": "", "description": "",
        "reason": "", "severity": "", "freetext": [], "extra": [],
    }
    known_keys = {
        "status":      r"^status\s*[:\-]\s*",
        "description": r"^description\s*[:\-]\s*",
        "reason":      r"^reason\s*[:\-]\s*",
        "severity":    r"^severity\s*[:\-]\s*",
    }

    lines          = [l.strip() for l in text.splitlines() if l.strip()]
    argos_stripped = False
    has_structured = False

    for line in lines:
        if not argos_stripped and line.lower() == ARGOS:
            fields["author"] = line
            argos_stripped   = True
            continue

        matched = False
        for field, pattern in known_keys.items():
            if re.match(pattern, line, re.IGNORECASE):
                fields[field] = re.sub(
                    pattern, "", line, flags=re.IGNORECASE).strip()
                matched        = True
                has_structured = True
                break

        if matched:
            continue

        if has_structured:
            fields["extra"].append(line)
        else:
            fields["freetext"].append(line)

    fields["_has_structured"] = has_structured
    return fields


def format_description(parsed: dict) -> str:
    has_structured = parsed.get("_has_structured", False)
    freetext       = parsed.get("freetext", [])
    extra          = parsed.get("extra",    [])

    if (not has_structured and not freetext
            and not parsed["status"] and not parsed["description"]
            and not parsed["reason"] and not parsed["severity"]):
        return "Strikethrough Text"

    if not has_structured and freetext:
        return "\n".join(freetext)

    lines = []
    if parsed["status"]:      lines.append(f"Status: {parsed['status']}")
    if parsed["description"]: lines.append(f"Description: {parsed['description']}")
    if parsed["reason"]:      lines.append(f"Reason: {parsed['reason']}")
    if parsed["severity"]:    lines.append(f"Severity: {parsed['severity']}")
    for ft in freetext:       lines.append(ft)
    for ex in extra:          lines.append(ex)

    return "\n".join(lines) if lines else "Strikethrough Text"


def build_keywords(parsed: dict) -> str:
    has_structured = parsed.get("_has_structured", False)
    freetext       = parsed.get("freetext", [])

    if (not has_structured and not freetext
            and not parsed["status"] and not parsed["description"]
            and not parsed["reason"] and not parsed["severity"]):
        return "N/A"

    keywords = []
    for word in parsed.get("severity", "").split():
        clean = re.sub(r'[^\w\-]', '', word)
        if clean and clean not in keywords:
            keywords.append(clean)

    all_text = " ".join([
        parsed.get("status", ""),
        parsed.get("description", ""),
        parsed.get("reason", ""),
        parsed.get("severity", ""),
        " ".join(parsed.get("freetext", [])),
    ])
    if re.search(r'\bDTP\b', all_text, re.IGNORECASE):
        if "DTP" not in keywords:
            keywords.insert(0, "DTP")

    return ", ".join(keywords) if keywords else "N/A"


def _detect_language(text: str) -> str:
    text = text.strip()
    if len(text) < 20:
        return "Unknown"
    try:
        return convert_to_bcp47(detect(text))
    except LangDetectException:
        return "Unknown"


def convert_to_bcp47(lang_code: str) -> str:
    lang_map = {
        "pl": "pl-PL", "en": "en-US", "de": "de-DE", "fr": "fr-FR",
        "es": "es-ES", "it": "it-IT", "pt": "pt-PT", "ru": "ru-RU",
        "nl": "nl-NL", "cs": "cs-CZ", "sk": "sk-SK", "hu": "hu-HU",
        "ro": "ro-RO", "bg": "bg-BG", "hr": "hr-HR", "sr": "sr-RS",
        "uk": "uk-UA", "sv": "sv-SE", "da": "da-DK", "fi": "fi-FI",
        "nb": "nb-NO", "tr": "tr-TR", "ar": "ar-SA",
        "zh-cn": "zh-CN", "zh-tw": "zh-TW",
        "ja": "ja-JP", "ko": "ko-KR",
    }
    return lang_map.get(lang_code.lower(), f"{lang_code}-{lang_code.upper()}")


def create_excel_report(comments: list[dict],
                         output_path: str,
                         status_callback=None) -> None:
    if status_callback:
        status_callback("  Building Excel workbook …")

    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Comments"
    wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    headers = [
        ("No.",                  8),
        ("File Name",           30),
        ("Description/Comment", 60),
        ("Found on Page",       15),
        ("Language",            15),
        ("Keywords Appearing",  30),
    ]

    for col_idx, (label, width) in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label)
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = 'A2'

    if not comments:
        ws.cell(row=2, column=1, value="No comments found in the PDF file.")
    else:
        for row_idx, comment in enumerate(comments, start=2):
            row_data = [
                comment["no"], comment["file_name"], comment["description"],
                comment["found_on_page"], comment["language"], comment["keywords"],
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = wrap
            line_count = comment["description"].count('\n') + 1
            ws.row_dimensions[row_idx].height = max(15, line_count * 15)

    ws_sum = wb.create_sheet("Summary")
    create_summary_sheet(ws_sum, comments)

    wb.save(output_path)


def create_summary_sheet(ws, comments: list[dict]) -> None:
    ws.cell(row=1, column=1, value="COMMENTS SUMMARY")
    ws.cell(row=3, column=1, value="Metric")
    ws.cell(row=3, column=2, value="Value")

    metrics = [
        ("Total Comments Found", len(comments)),
        ("Pages with Comments",
         len(set(c["found_on_page"] for c in comments)) if comments else 0),
        ("Languages Detected",
         len(set(c["language"]     for c in comments)) if comments else 0),
    ]
    for idx, (metric, value) in enumerate(metrics, start=4):
        ws.cell(row=idx, column=1, value=metric)
        ws.cell(row=idx, column=2, value=value)

    if not comments:
        return

    lang_ctr = Counter(c["language"] for c in comments)
    kw_ctr   = Counter(
        kw.strip()
        for c in comments
        for kw in c["keywords"].split(", ")
        if kw.strip() and kw.strip() != "N/A"
    )

    lang_start = 4 + len(metrics) + 1
    ws.cell(row=lang_start,     column=1, value="Language Breakdown")
    ws.cell(row=lang_start + 1, column=1, value="Language")
    ws.cell(row=lang_start + 1, column=2, value="Count")
    for idx, (lang, count) in enumerate(lang_ctr.most_common(),
                                        start=lang_start + 2):
        ws.cell(row=idx, column=1, value=lang)
        ws.cell(row=idx, column=2, value=count)

    kw_start = lang_start + 2 + len(lang_ctr) + 1
    ws.cell(row=kw_start,     column=1, value="Keywords Breakdown")
    ws.cell(row=kw_start + 1, column=1, value="Keyword")
    ws.cell(row=kw_start + 1, column=2, value="Occurrences")
    for idx, (kw, count) in enumerate(kw_ctr.most_common(),
                                      start=kw_start + 2):
        ws.cell(row=idx, column=1, value=kw)
        ws.cell(row=idx, column=2, value=count)

def unique_output_path(path: str) -> str:
    p = Path(path)
    if not p.exists():
        return path
    stem    = p.stem
    suffix  = p.suffix
    parent  = p.parent
    counter = 1
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return str(candidate)
        counter += 1

def collect_pdfs_from_folder(folder: str,
                              recursive: bool = True
                              ) -> tuple[list[str], list[str]]:
    root      = Path(folder)
    pdf_paths: list[str] = []
    temp_dirs: list[str] = []
    pattern   = "**/*" if recursive else "*"

    for item in sorted(root.glob(pattern)):
        if not item.is_file():
            continue
        if item.suffix.lower() == ".pdf":
            pdf_paths.append(str(item.resolve()))
        elif item.suffix.lower() == ".zip":
            tmp = tempfile.mkdtemp(prefix="pdf_extractor_zip_")
            temp_dirs.append(tmp)
            try:
                with zipfile.ZipFile(str(item), "r") as zf:
                    zf.extractall(tmp)
                inner_pdfs, inner_tmps = collect_pdfs_from_folder(
                    tmp, recursive=True)
                pdf_paths.extend(inner_pdfs)
                temp_dirs.extend(inner_tmps)
            except Exception:
                pass

    return pdf_paths, temp_dirs


def extract_pdfs_from_zip(zip_path: str) -> tuple[list[str], list[str], list[str]]:
    tmp = tempfile.mkdtemp(prefix="pdf_extractor_zip_")
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            members = zf.namelist()
            zf.extractall(tmp)
    except Exception as exc:
        shutil.rmtree(tmp, ignore_errors=True)
        raise RuntimeError(
            f"Cannot open ZIP '{Path(zip_path).name}': {exc}") from exc

    pdfs, temps = collect_pdfs_from_folder(tmp, recursive=True)
    temps.append(tmp)
    return pdfs, temps, members

class SourceEntry:
    __slots__ = ("kind", "origin_path", "output_root", "pdf_paths")

    def __init__(self, kind: str, origin_path: str,
                 output_root: str, pdf_paths: list[str]):
        self.kind        = kind
        self.origin_path = origin_path
        self.output_root = output_root   
        self.pdf_paths   = pdf_paths

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF Comment Extractor")
        self.resizable(False, False)
        self.configure(padx=24, pady=20, bg="#f5f5f5")

        self._sources:   list[SourceEntry] = []
        self._pdf_paths: list[str]         = []
        self._temp_dirs: list[str]         = []
        self._output_dir = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        BG     = "#f5f5f5"
        ACCENT = "#2563eb"
        FONT   = ("Segoe UI", 10)
        FONT_B = ("Segoe UI", 10, "bold")
        FONT_H = ("Segoe UI", 14, "bold")

        row = 0

        tk.Label(self, text="PDF Comment Extractor",
                 font=FONT_H, bg=BG, fg="#1e293b").grid(
            row=row, column=0, columnspan=3,
            pady=(0, 18), sticky="w")
        row += 1

        tk.Label(self, text="Input PDFs", font=FONT_B,
                 bg=BG, fg="#334155").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(0, 4))
        row += 1

        list_frame = tk.Frame(self, bg=BG)
        list_frame.grid(row=row, column=0, columnspan=3,
                        sticky="nsew", pady=(0, 4))

        self._pdf_listbox = tk.Listbox(
            list_frame,
            height=7, width=62,
            selectmode=tk.EXTENDED,
            font=("Consolas", 9),
            bg="#1e293b", fg="#e2e8f0",
            selectbackground="#2563eb",
            selectforeground="white",
            relief="flat",
            activestyle="none",
        )
        lb_scroll = ttk.Scrollbar(list_frame, command=self._pdf_listbox.yview)
        self._pdf_listbox.configure(yscrollcommand=lb_scroll.set)
        self._pdf_listbox.pack(side="left", fill="both", expand=True)
        lb_scroll.pack(side="right", fill="y")
        row += 1

        self._list_hint_var = tk.StringVar(value="No PDFs selected.")
        tk.Label(self, textvariable=self._list_hint_var,
                 font=("Segoe UI", 8), bg=BG, fg="#94a3b8",
                 anchor="w").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(0, 6))
        row += 1

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.grid(row=row, column=0, columnspan=3,
                       sticky="w", pady=(0, 14))

        btn_cfg = dict(font=FONT, bg="#e2e8f0", fg="#1e293b",
                       relief="flat", activebackground="#cbd5e1",
                       cursor="hand2", padx=10, pady=4)

        self._btn_add_pdfs = tk.Button(
            btn_frame, text="＋ Add PDFs …",
            command=self._add_pdfs, **btn_cfg)
        self._btn_add_pdfs.pack(side="left", padx=(0, 6))

        self._btn_add_zip = tk.Button(
            btn_frame, text="📦 Add ZIP …",
            command=self._add_zip, **btn_cfg)
        self._btn_add_zip.pack(side="left", padx=(0, 6))

        self._btn_add_folder = tk.Button(
            btn_frame, text="📁 Add Folder …",
            command=self._add_folder, **btn_cfg)
        self._btn_add_folder.pack(side="left", padx=(0, 6))

        self._btn_remove = tk.Button(
            btn_frame, text="✕ Remove Selected",
            command=self._remove_selected_pdfs, **btn_cfg)
        self._btn_remove.pack(side="left", padx=(0, 6))

        self._btn_clear = tk.Button(
            btn_frame, text="🗑 Clear All",
            command=self._clear_pdfs, **btn_cfg)
        self._btn_clear.pack(side="left")
        row += 1

        tk.Label(self, text="Output Folder", font=FONT_B,
                 bg=BG, fg="#334155").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(0, 4))
        row += 1

        self._dir_name_var = tk.StringVar(
            value="Auto  (next to each source file / folder)")
        tk.Label(self, textvariable=self._dir_name_var,
                 font=("Segoe UI", 9), bg="#e2e8f0", fg="#1e293b",
                 anchor="w", relief="solid", bd=1, padx=6, pady=4,
                 width=46).grid(
            row=row, column=0, columnspan=2, sticky="ew")

        self._btn_browse_out = tk.Button(
            self, text="Browse …", font=FONT,
            bg="#e2e8f0", fg="#1e293b", relief="flat",
            activebackground="#cbd5e1", cursor="hand2",
            command=self._browse_output_dir)
        self._btn_browse_out.grid(
            row=row, column=2, padx=(8, 0), ipady=4, ipadx=6)
        row += 1

        self._btn_clear_out = tk.Button(
            self, text="✕ Reset to auto",
            font=("Segoe UI", 8), bg=BG, fg="#94a3b8",
            relief="flat", cursor="hand2",
            activebackground=BG, activeforeground="#64748b",
            command=self._clear_output_dir)
        self._btn_clear_out.grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(2, 0))
        row += 1

        tk.Label(self, text="Progress", font=FONT_B,
                 bg=BG, fg="#334155").grid(
            row=row, column=0, sticky="w", pady=(14, 4))
        row += 1

        self._progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(self, variable=self._progress_var,
                        maximum=100, length=460,
                        mode="determinate").grid(
            row=row, column=0, columnspan=3, sticky="ew")
        row += 1

        self._status_var = tk.StringVar(value="Ready.")
        tk.Label(self, textvariable=self._status_var,
                 font=("Segoe UI", 9), bg=BG, fg="#64748b",
                 anchor="w", width=60).grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(6, 0))
        row += 1

        tk.Label(self, text="Log", font=FONT_B,
                 bg=BG, fg="#334155").grid(
            row=row, column=0, sticky="w", pady=(14, 4))
        row += 1

        log_frame = tk.Frame(self, bg=BG)
        log_frame.grid(row=row, column=0, columnspan=3, sticky="nsew")

        self._log = tk.Text(
            log_frame, height=10, width=62,
            font=("Consolas", 9),
            bg="#1e293b", fg="#e2e8f0",
            relief="flat", state="disabled",
            wrap="word", padx=8, pady=6)
        scrollbar = ttk.Scrollbar(log_frame, command=self._log.yview)
        self._log.configure(yscrollcommand=scrollbar.set)
        self._log.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        row += 1

        self._run_btn = tk.Button(
            self, text="▶  Run Extraction",
            font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg="white",
            activebackground="#1d4ed8", activeforeground="white",
            relief="flat", cursor="hand2",
            command=self._start_extraction,
            pady=8)
        self._run_btn.grid(
            row=row, column=0, columnspan=3,
            sticky="ew", pady=(18, 0))

    def _browse_output_dir(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self._output_dir.set(folder)
            self._dir_name_var.set(str(Path(folder)))

    def _clear_output_dir(self):
        self._output_dir.set("")
        self._dir_name_var.set(
            "Auto  (next to each source file / folder)")

    def _resolve_output_path(self, pdf_path: str) -> str:
        out_name = Path(pdf_path).stem + ".xlsx"
        global_dir = self._output_dir.get().strip()

        if global_dir:
            out_folder = Path(global_dir)
        else:
            source = self._find_source_for_pdf(pdf_path)
            if source is None:
                out_folder = Path(pdf_path).parent
            else:
                out_folder = Path(source.output_root)

        out_folder.mkdir(parents=True, exist_ok=True)
        return unique_output_path(str(out_folder / out_name))

    def _find_source_for_pdf(self, pdf_path: str) -> "SourceEntry | None":
        for src in self._sources:
            if pdf_path in src.pdf_paths:
                return src
        return None

    @staticmethod
    def _output_root_for_zip(zip_path: str) -> str:
        z = Path(zip_path).resolve()
        return str(z.parent / f"{z.stem} - output")

    @staticmethod
    def _output_root_for_folder(folder_path: str) -> str:
        f = Path(folder_path).resolve()
        return str(f.parent / f"{f.name} - output")

    @staticmethod
    def _output_root_for_file(pdf_path: str) -> str:
        return str(Path(pdf_path).resolve().parent)

    def _add_pdfs(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if not paths:
            return

        added = 0
        for raw in paths:
            p = str(Path(raw).resolve())
            if p in self._pdf_paths:
                continue
            src = SourceEntry(
                kind        = "file",
                origin_path = p,
                output_root = self._output_root_for_file(p),
                pdf_paths   = [p],
            )
            self._sources.append(src)
            self._pdf_paths.append(p)
            added += 1

        self._log_line(f"Added {added} PDF(s) via file picker.")
        self._refresh_pdf_list_ui()

    def _add_zip(self):
        zip_path = filedialog.askopenfilename(
            title="Select a ZIP containing PDFs",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")])
        if not zip_path:
            return

        zip_path = str(Path(zip_path).resolve())
        zip_name = Path(zip_path).name

        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                all_members = zf.namelist()
                pdf_members = [n for n in all_members
                               if n.lower().endswith(".pdf")]
                zip_members = [n for n in all_members
                               if n.lower().endswith(".zip")]
        except zipfile.BadZipFile as exc:
            messagebox.showerror("ZIP Error",
                                 f"Not a valid ZIP:\n{zip_name}\n\n{exc}")
            return
        except Exception as exc:
            messagebox.showerror("ZIP Error",
                                 f"Cannot read ZIP:\n{zip_name}\n\n{exc}")
            return

        self._log_line(f"📦 ZIP '{zip_name}': "
                       f"{len(pdf_members)} direct PDF(s), "
                       f"{len(zip_members)} nested ZIP(s).")

        try:
            pdfs, temps, _ = extract_pdfs_from_zip(zip_path)
        except RuntimeError as exc:
            messagebox.showerror("ZIP Error", str(exc))
            return

        if not pdfs:
            messagebox.showwarning(
                "No PDFs found",
                f"'{zip_name}' contained no PDF files.")
            for d in temps:
                shutil.rmtree(d, ignore_errors=True)
            return

        self._temp_dirs.extend(temps)

        output_root = self._output_root_for_zip(zip_path)
        self._log_line(
            f"   Output folder will be: {output_root}")

        new_pdfs = [p for p in pdfs if p not in self._pdf_paths]
        src = SourceEntry(
            kind        = "zip",
            origin_path = zip_path,
            output_root = output_root,
            pdf_paths   = new_pdfs,
        )
        self._sources.append(src)
        self._pdf_paths.extend(new_pdfs)

        self._log_line(
            f"   Added {len(new_pdfs)} PDF(s) from '{zip_name}'.")
        for p in new_pdfs:
            self._log_line(f"     {Path(p).name}")
        self._refresh_pdf_list_ui()

    def _add_folder(self):
        folder = filedialog.askdirectory(
            title="Select folder (scanned recursively)")
        if not folder:
            return

        folder = str(Path(folder).resolve())
        self._log_line(f"📁 Scanning folder: {folder}")

        pdfs, temps = collect_pdfs_from_folder(folder, recursive=True)
        self._temp_dirs.extend(temps)

        output_root = self._output_root_for_folder(folder)
        self._log_line(
            f"   Output folder will be: {output_root}")

        new_pdfs = [p for p in pdfs if p not in self._pdf_paths]
        src = SourceEntry(
            kind        = "folder",
            origin_path = folder,
            output_root = output_root,
            pdf_paths   = new_pdfs,
        )
        self._sources.append(src)
        self._pdf_paths.extend(new_pdfs)

        self._log_line(
            f"   Found {len(pdfs)} PDF(s), "
            f"added {len(new_pdfs)} new.")
        self._refresh_pdf_list_ui()

    def _remove_selected_pdfs(self):
        selected = list(self._pdf_listbox.curselection())
        if not selected:
            return
        to_remove = {self._pdf_paths[i] for i in selected}
        self._pdf_paths = [p for p in self._pdf_paths
                           if p not in to_remove]
        for src in self._sources:
            src.pdf_paths = [p for p in src.pdf_paths
                             if p not in to_remove]
        self._sources = [s for s in self._sources if s.pdf_paths]
        self._refresh_pdf_list_ui()

    def _clear_pdfs(self):
        self._sources.clear()
        self._pdf_paths.clear()
        self._cleanup_temp_dirs()
        self._refresh_pdf_list_ui()

    def _refresh_pdf_list_ui(self):
        self._pdf_listbox.delete(0, tk.END)
        for p in self._pdf_paths:
            src = self._find_source_for_pdf(p)
            if src and src.kind in ("zip", "folder"):
                tag = f"[{src.kind.upper()}: {Path(src.origin_path).name}]"
            else:
                tag = f"[{Path(p).parent}]"
            display = f"{Path(p).name}   {tag}"
            self._pdf_listbox.insert(tk.END, display)

        count = len(self._pdf_paths)
        if count == 0:
            self._list_hint_var.set("No PDFs selected.")
        else:
            self._list_hint_var.set(
                f"{count} PDF(s) queued  —  "
                "Ctrl+click / Shift+click to multi-select for removal.")

    def _log_line(self, message: str):
        def _do():
            self._log.configure(state="normal")
            self._log.insert("end", message + "\n")
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _do)

    def _set_status(self, message: str):
        def _do():
            self._status_var.set(message)
            self.update_idletasks()
        self.after(0, _do)
        self._log_line(message)

    def _set_progress(self, value: float):
        def _do():
            self._progress_var.set(value)
            self.update_idletasks()
        self.after(0, _do)

    def _set_controls_state(self, state: str):
        for btn in (self._btn_add_pdfs, self._btn_add_zip,
                    self._btn_add_folder, self._btn_remove,
                    self._btn_clear, self._btn_browse_out,
                    self._btn_clear_out, self._run_btn):
            btn.configure(state=state)

    def _start_extraction(self):
        if not self._pdf_paths:
            messagebox.showwarning(
                "Nothing to process",
                "Please add at least one PDF file before running.")
            return

        missing = [p for p in self._pdf_paths if not Path(p).exists()]
        if missing:
            answer = messagebox.askyesno(
                "Missing files",
                f"{len(missing)} file(s) no longer exist on disk "
                f"and will be skipped.\n\nContinue with the rest?")
            if not answer:
                return
            for p in missing:
                self._pdf_paths.remove(p)
                for src in self._sources:
                    src.pdf_paths = [x for x in src.pdf_paths if x != p]
            self._sources = [s for s in self._sources if s.pdf_paths]
            self._refresh_pdf_list_ui()
            if not self._pdf_paths:
                return

        self._progress_var.set(0)
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        self._set_controls_state("disabled")

        self._log_line(
            f"Starting batch: {len(self._pdf_paths)} PDF(s) queued.")
        for i, p in enumerate(self._pdf_paths, 1):
            src = self._find_source_for_pdf(p)
            src_label = (f"{src.kind.upper()}: "
                         f"{Path(src.origin_path).name}"
                         if src else "?")
            self._log_line(f"  [{i}] {Path(p).name}  ({src_label})")
        self._log_line("─" * 50)

        thread = threading.Thread(
            target=self._run_batch_extraction,
            args=(self._pdf_paths.copy(),),
            daemon=True)
        thread.start()

    def _run_batch_extraction(self, pdf_paths: list[str]):
        total    = len(pdf_paths)
        success  = 0
        failed   = 0
        failures: list[tuple[str, str]] = []

        try:
            for file_index, pdf_path in enumerate(pdf_paths):
                fname = Path(pdf_path).name
                self._set_status(f"[{file_index + 1}/{total}] {fname}")
                self._log_line("─" * 50)
                self._log_line(f"▶ [{file_index + 1}/{total}] {fname}")

                output_path = self._resolve_output_path(pdf_path)
                self._log_line(f"   → {output_path}")

                try:
                    def make_progress_cb(idx: int):
                        def _cb(p: float):
                            overall = ((idx + p / 100.0) / total) * 100
                            self._set_progress(overall)
                        return _cb

                    comments = extract_comments_from_pdf(
                        pdf_path,
                        progress_callback=make_progress_cb(file_index),
                        status_callback=self._set_status,
                    )
                    self._log_line(
                        f"   {len(comments)} comment(s) found.")

                    create_excel_report(
                        comments,
                        output_path,
                        status_callback=self._set_status,
                    )
                    self._log_line(
                        f"  ✓ Saved → {Path(output_path).name}")
                    success += 1

                except Exception as exc:
                    import traceback
                    self._log_line(f"  ✗ FAILED: {exc}")
                    self._log_line(traceback.format_exc())
                    failures.append((fname, str(exc)))
                    failed += 1

            self._set_progress(100)
            self._log_line("─" * 50)
            self._log_line(
                f"✓ Batch done.  "
                f"Total: {total}  Success: {success}  Failed: {failed}")

            global_dir = self._output_dir.get().strip()
            if global_dir:
                folder_label = global_dir
            else:
                roots = list(dict.fromkeys(
                    s.output_root for s in self._sources))
                folder_label = "\n".join(roots) if roots else "see log"

            summary = (
                f"Batch extraction complete!\n\n"
                f"  PDFs selected : {total}\n"
                f"  Succeeded     : {success}\n"
                f"  Failed        : {failed}\n\n"
                f"Output location(s):\n{folder_label}"
            )
            if failures:
                detail = "\n".join(
                    f"  • {n}: {m}" for n, m in failures)
                summary += f"\n\nFailed files:\n{detail}"

            self.after(0, lambda: messagebox.showinfo(
                "Batch Complete", summary))

        finally:
            self._cleanup_temp_dirs()
            self.after(0, self._reset_controls)

    def _cleanup_temp_dirs(self):
        for d in self._temp_dirs:
            shutil.rmtree(d, ignore_errors=True)
        self._temp_dirs.clear()

    def _reset_controls(self):
        self._set_controls_state("normal")
        self._status_var.set("Ready.")

if __name__ == "__main__":
    app = App()
    app.mainloop()